from aiogram import types

from data.config import admin_ids
from db import ShopsData
from handlers.users.help import CheckOrBuyBTNs
from loader import dp

from openpyxl import load_workbook


@dp.message_handler(content_types=types.ContentTypes.DOCUMENT)
@dp.message_handler(content_types="document")
async def upload_db_shops(m: types.Message):
    print('загрузка магазов')
    if m.from_user.id in admin_ids:
        load_message = await m.answer(
            "Началась загрузка базы артикулов...",
        )
        try:
            await ShopsData.all().delete()
            file = await dp.bot.get_file(m.document.file_id)
            await dp.bot.download_file(file.file_path, "base.xlsx")
            wb = load_workbook("base.xlsx")
            sheet = wb.active
            for cells in sheet.iter_rows():
                item = [cell for cell in cells]

                await ShopsData.create(city=item[0].value,
                                       district=item[1].value,
                                       street=item[2].value,
                                       link=f"https://t.me/{str(item[3].value)[1:]}",
                                       link_card_ya="https://yandex.ru/maps/-/CDDx5Xmy",
                                       link_card_2gis="https://go.2gis.com/syq0a")

            await dp.bot.edit_message_text(
                "Загрузка базы успешно завершена",
                chat_id=m.chat.id,
                message_id=load_message.message_id
            )
        except Exception as e:
            await dp.bot.edit_message_text(
                "При загрузке меню произошла ошибка. Проверьте структуру файла и повторите попытку.",
                chat_id=m.chat.id,
                message_id=load_message.message_id
            )
            await m.answer(f'{e}')


@dp.callback_query_handler(text_startswith="check_or_buy")
async def check_or_buy(c: types.CallbackQuery):
    m_id = str(c.data).split(":")[1]
    await dp.bot.edit_message_text(text="🤔 Давай выберем город:",
                                   chat_id=c.from_user.id,
                                   message_id=m_id,
                                   reply_markup=await CheckOrBuyBTNs.get_cities(m_id=m_id))


@dp.callback_query_handler(text_startswith="city")
async def get_city_take_district(c: types.CallbackQuery):
    city = str(c.data).split(":")[1]
    m_id = str(c.data).split(":")[2]
    await dp.bot.edit_message_text(text="🤞🏻Отлично! А теперь давай выберем район или улицу:",
                                   chat_id=c.from_user.id,
                                   message_id=m_id,
                                   reply_markup=await CheckOrBuyBTNs.get_district(city=city, m_id=m_id))


@dp.callback_query_handler(text_startswith="district")
async def get_district_take_street(c: types.CallbackQuery):
    district = str(c.data).split(":")[1]
    m_id = str(c.data).split(":")[2]
    await dp.bot.edit_message_text(text="✴️Какой магазин будет ближе всех?",
                                   chat_id=c.from_user.id,
                                   message_id=m_id,
                                   reply_markup=await CheckOrBuyBTNs.get_street(district=district, m_id=m_id))


@dp.callback_query_handler(text_startswith="street")
async def get_street_take_links(c: types.CallbackQuery):
    street = str(c.data).split(":")[1]
    m_id = str(c.data).split(":")[2]
    print(street)
    await dp.bot.edit_message_text(text=f"<b>{street}</b>\n\n"
                                        f"Написать и спросить о наличии можно ниже, также я дам вам пару ссылок для "
                                        f"поиска магазина на картах",
                                   chat_id=c.from_user.id,
                                   message_id=m_id,
                                   reply_markup=await CheckOrBuyBTNs.get_links(street=street, m_id=m_id))
